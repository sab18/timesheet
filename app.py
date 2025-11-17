import tkinter as tk
from tkinter import ttk
from tkcalendar import Calendar
from openpyxl import Workbook, load_workbook
import pandas as pd
import yaml
from datetime import datetime, timedelta
import math
import json
import os
from tkinter import filedialog
from collections import defaultdict

from inputs import (yaml_file_path_projects, 
                    yaml_file_path_locations, 
                    excel_file_path, 

                    excel_sheet_names,

                    headers_default_data_for_timesheet,
                    headers_default_data_daily_summaries)

class ScrollableFrame(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        canvas = tk.Canvas(self, borderwidth=0, background="#e0ffe0")
        scrollbar = tk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, background="#e0ffe0")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        canvas.columnconfigure(0, weight=1)
        canvas.rowconfigure(0, weight=1)
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Only bind mouse wheel when cursor is over this canvas
        canvas.bind("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1*(event.delta/120)), "units"))
        

class Timesheet(tk.Tk):

    def __init__(self):
        super().__init__()

        self.title("Timesheet")
       
        window_width = 800
        window_height = 600
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        self.geometry(f"{window_width}x{window_height}+{x}+{y-35}")


        self.font = "Cascadia Code"
        self.font_header = (self.font, 12, "bold")
        self.font_normal = (self.font, 10)
        self.padding={"padx":10, "pady":10}
        self.sticky_w = {"sticky":"w"}
        self.sticky_e = {"sticky":"e"}
        self.sticky_ew = {"sticky":"ew"}
        self.sticky_all = {"sticky":"nsew"}
        



        # self.dict_projects = self.import_yaml(yaml_file_path_projects)
        # self.dict_projects.pop("Retired", None)
        # self.dict_locations = self.import_yaml(yaml_file_path_locations)

        config = self.load_config()
        self.locations_yaml_path = config.get("locations_yaml", "locations.yaml")
        self.projects_yaml_path = config.get("projects_yaml", "projects.yaml")
        self.timesheet_excel_path = config.get("timesheet_excel", "Timesheet_2025.xlsx")

        self.dict_projects = self.import_yaml(self.projects_yaml_path)
        if "Retired" in self.dict_projects:
                self.dict_projects.pop("Retired")
        
        self.dict_locations = self.import_yaml(self.locations_yaml_path)



        self.fetch_excel_data()

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)
        self.notebook = ttk.Notebook(self)
        self.notebook.grid(row=0, column=0, sticky="nsew")
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)


        
        self.today = datetime.now()

        self.tab_input()
        self.tab_display()
        self.tab_monthly_fte()
        self.tab_yearly_fte()
        self.tab_config()

    def import_yaml(self, file_path):
    
        with open(file_path, "r") as file:
           dict = yaml.safe_load(file)

            
        return dict
    

    def fetch_excel_data(self):
        self.timesheet_wb = load_workbook(self.timesheet_excel_path)
        self.timesheet_ws_hrs = self.timesheet_wb[excel_sheet_names["for_timesheet"]]
        self.timesheet_ws_hrs_column_index = self.get_columns_by_header(self.timesheet_ws_hrs, headers_default_data_for_timesheet)
        self.timesheet_ws_daily_summaries = self.timesheet_wb[excel_sheet_names["daily_summaries"]]
        self.timesheet_ws_daily_summaries_column_index = self.get_columns_by_header(self.timesheet_ws_daily_summaries, headers_default_data_daily_summaries)
        
    def load_config(self):
        config_path = "config.json"
        if os.path.exists(config_path):
            with open(config_path, "r") as f:
                return json.load(f)
        return {}

    def save_config(self):
        config_path = "config.json"
        config = {
            "locations_yaml": self.locations_yaml_path,
            "projects_yaml": self.projects_yaml_path,
            "timesheet_excel": self.timesheet_excel_path
        }
        with open(config_path, "w") as f:
            json.dump(config, f)


    def tab_input(self):
        self.tab_input_frame = tk.Frame(self.notebook, bg="lightblue")
        self.notebook.add(self.tab_input_frame, text="Input Hrs")
        self.tab_input_frame.rowconfigure(0, weight=1)
        self.tab_input_frame.columnconfigure(0, weight=1)
        self.tab_input_frame.columnconfigure(1, weight=1)

        self.section_input_hrs()
        self.section_input_details()

    def section_input_hrs(self):
        hrs_scroll_frame = ScrollableFrame(self.tab_input_frame)
        hrs_scroll_frame.grid(row=0, column=0, **self.sticky_all)
        hrs_scroll_frame.scrollable_frame.columnconfigure(0, weight=1)
        hrs_scroll_frame.scrollable_frame.columnconfigure(1, weight=1)
        hrs_scroll_frame.grid_rowconfigure(0, weight=1)
        hrs_scroll_frame.grid_columnconfigure(0, weight=1)

     
        vcmd = (self.register(self.is_valid_positive_number), "%P")

        row_num=0
        
        self.entries = {}
        # print(self.dict_projects)

        for category in self.dict_projects:

            
            self.label = tk.Label( hrs_scroll_frame.scrollable_frame, text=category, font=self.font_header)
            self.label.grid(row=row_num, column=0, **self.padding, **self.sticky_e)
            row_num += 1

            for project in self.dict_projects[category]:
                project_nickname = project['project_nickname']
                project_name = project['project_name']
                cost_center_name = project['cost_center_name']
                cost_center_code = project['cost_center_code']
           
            
                self.label = tk.Label( hrs_scroll_frame.scrollable_frame, text=f"  {project_nickname}", font=self.font_normal)
                self.label.grid(row=row_num, column=0, **self.padding, **self.sticky_e)

                entry = tk.Entry(hrs_scroll_frame.scrollable_frame, font=self.font_normal, width=10 , validate="key", validatecommand=vcmd)
                entry.grid(row=row_num, column=1, **self.padding, **self.sticky_w)
                
                row_num += 1
                self.entries[project_nickname] = (project_name, cost_center_name, cost_center_code, entry)
    
    def section_input_details(self):
        section_input_details = tk.Frame(self.tab_input_frame, bg="lightyellow")
        section_input_details.grid(row=0, column=1, **self.sticky_all)
        section_input_details.rowconfigure(0, weight=1)
        section_input_details.columnconfigure(0, weight=1)

        i=0

        self.calendar = Calendar(section_input_details, selectmode="day", firstweekday='sunday', showweeknumbers=False, year=self.today.year, month=self.today.month, day=self.today.day)
        self.calendar.grid(row=0, column=0, **self.padding)

        self.label_location = tk.Label(section_input_details, text="Location:", font=self.font_normal)
        self.label_location.grid(row=i+1, column=0, **self.padding, **self.sticky_w)

        self.combobox_location = ttk.Combobox(section_input_details, values=self.dict_locations, font=self.font_normal)
        self.combobox_location.grid(row=i+2, column=0, **self.padding, **self.sticky_w)

        self.label_summary = tk.Label(section_input_details, text="What I did today:", font=self.font_normal)
        self.label_summary.grid(row=i+3, column=0, **self.padding, **self.sticky_w)

        self.text_summary = tk.Text(section_input_details, font=self.font_normal, wrap=tk.WORD, width=50, height=5)
        self.text_summary.grid(row=i+4, column=0, **self.padding, **self.sticky_ew)

        self.button_save = tk.Button(section_input_details, text="Save", font=self.font_normal, command=self.section_input_button_save)
        self.button_save.grid(row=i+5, column=0, **self.padding, **self.sticky_e)

        self.submission_status = tk.Label(section_input_details, text="", font=self.font_normal)
        self.submission_status.grid(row=i+6, column=0, **self.padding, **self.sticky_e)
    

    def fetch_all_inputs(self):

        
        hours_value = {}
        for project_nickname, (project_name, cost_center_name, cost_center_code, entry) in self.entries.items():
            value = entry.get()
            hours_value[project_nickname] = (project_name, cost_center_name, cost_center_code, value)

        calendar_value = self.calendar.get_date()
        combobox_location_value = self.combobox_location.get()
        text_summary_value = self.text_summary.get("1.0", "end-1c")

        return hours_value, calendar_value, combobox_location_value, text_summary_value

    def section_input_button_save(self):
        
        hours_value, calendar_value, combobox_location_value, text_summary_value = self.fetch_all_inputs()

        #try not to change column names bc it'll mess everything up!
        
        next_row_hrs = self.timesheet_ws_hrs.max_row + 1
        next_row_daily_summaries = self.timesheet_ws_daily_summaries.max_row + 1
        

        for key, value in hours_value.items():
            # print(key, value)
            if value[3] not in ("", "0", 0):
                
                self.timesheet_ws_hrs.cell(row=next_row_hrs, column=self.timesheet_ws_hrs_column_index["date"], value=calendar_value)
                self.timesheet_ws_hrs.cell(row=next_row_hrs, column=self.timesheet_ws_hrs_column_index["project_nickname"], value=key)
                self.timesheet_ws_hrs.cell(row=next_row_hrs, column=self.timesheet_ws_hrs_column_index["project_name"], value=value[1])
                self.timesheet_ws_hrs.cell(row=next_row_hrs, column=self.timesheet_ws_hrs_column_index["cost_center_name"], value=value[1])
                self.timesheet_ws_hrs.cell(row=next_row_hrs, column=self.timesheet_ws_hrs_column_index["cost_center_code"], value=value[2])
                self.timesheet_ws_hrs.cell(row=next_row_hrs, column=self.timesheet_ws_hrs_column_index["hrs"], value=value[3])

                next_row_hrs += 1
        
        self.timesheet_ws_daily_summaries.cell(row=next_row_daily_summaries, column=self.timesheet_ws_daily_summaries_column_index["date"], value=calendar_value)
        self.timesheet_ws_daily_summaries.cell(row=next_row_daily_summaries, column=self.timesheet_ws_daily_summaries_column_index["location"], value=combobox_location_value)
        self.timesheet_ws_daily_summaries.cell(row=next_row_daily_summaries, column=self.timesheet_ws_daily_summaries_column_index["note"], value=text_summary_value)

        if combobox_location_value != "" and text_summary_value != "":
            self.timesheet_wb.save(excel_file_path)

            self.clear_entries()
            self.fetch_excel_data()
            self.show_temporary_message(self.submission_status, message="submitted!")
            
        else:
            self.show_temporary_message(self.submission_status, message="missing field!")

    

    def get_columns_by_header(self, worksheet, headers, header_row=1):
        dict_columns_num = {}
     
        for header in headers:

            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col).value
                if cell_value == header:
                    dict_columns_num[header] = col
        return dict_columns_num

    def clear_entries(self):

        for project_nickname, (project_name, cost_center_name, cost_center_code, entry) in self.entries.items():
            entry.delete(0, tk.END)
            
        self.combobox_location.set("")
        self.text_summary.delete("1.0", tk.END)


    def show_temporary_message(self, label, message, duration=2000):
        label.config(text=message)
        label.after(duration, lambda: label.config(text=""))
        
    def tab_display(self):
        tab_display = tk.Frame(self.notebook)
        self.notebook.add(tab_display, text="View Hrs")
        tab_display.columnconfigure(0, weight=1)
        tab_display.rowconfigure(0, weight=1)

        # Create canvas and scrollbar
        canvas = tk.Canvas(tab_display)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        scrollbar = ttk.Scrollbar(tab_display, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Create a frame inside the canvas
        self.display_frame = tk.Frame(canvas)
        self.display_frame.columnconfigure(0, weight=1)
        self.display_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.display_frame, anchor="nw")

        self.week_tables = []
        for i in range(4):
            week_frame = tk.Frame(self.display_frame)
            week_frame.grid(row=i, column=0, sticky="nsew", pady=10, padx=(0,15))
            week_frame.columnconfigure(0, weight=1)
            week_frame.columnconfigure(1, weight=1)
            self.display_frame.rowconfigure(i, weight=1)

            self.week_tables.append(week_frame)
        self.display_timesheet_data()

    def display_timesheet_data(self):
        
        # Get all timesheet data
        all_data = []
        for row in self.timesheet_ws_hrs.iter_rows(min_row=2, values_only=True):
            date = row[self.timesheet_ws_hrs_column_index["date"] - 1]
            cost_center_name = row[self.timesheet_ws_hrs_column_index["cost_center_name"] - 1]
            hrs = row[self.timesheet_ws_hrs_column_index["hrs"] - 1]
            if date and cost_center_name and hrs:
                try:
                    hrs_val = float(hrs)
                except ValueError:
                    continue
                all_data.append((date, cost_center_name, hrs_val))

        # Find last 4 Mondays
        today = self.today
        mondays = []
        # Go back to most recent Monday
        offset = (today.weekday() - 0) % 7
        most_recent_monday = today - timedelta(days=offset)
        for i in range(4):
            mondays.append(most_recent_monday - timedelta(weeks=i))
        mondays = sorted(mondays)

        # For each week, build M-F dates
        week_dates = []
        for monday in mondays:
            week = [(monday + timedelta(days=d)).date() for d in range(5)]
            week_dates.append(week)

        # Aggregate data for each week
        for week_idx, dates in enumerate(reversed(week_dates)):
            # Build cost_center_name -> {date: sum_hrs}
            agg = {}
            date_set = set(dates)
            for date, cost_center, hrs in all_data:
                # Convert date to date object if it's string
                if isinstance(date, str):
                    try:
                        date_obj = pd.to_datetime(date).date()
                    except Exception:
                        continue
                else:
                    date_obj = date if hasattr(date, 'date') else date
                if date_obj in date_set:
                    agg.setdefault(cost_center, {}).setdefault(date_obj, 0)
                    agg[cost_center][date_obj] += hrs
            # Normalize agg so each day's total is 8 hours
            normalized_agg = {}
            for d in dates:
                # Get original values for the day
                orig_values = {cc: agg.get(cc, {}).get(d, 0) for cc in agg}
                total = sum(orig_values.values())
                if total == 0:
                    # Leave all at 0
                    for cc in agg:
                        if cc not in normalized_agg:
                            normalized_agg[cc] = {}
                        normalized_agg[cc][d] = 0
                else:
                    # Scale to 8
                    scale = 8 / total
                    rounded_values = {cc: rounding(val * scale) for cc, val in orig_values.items()}
                    rounded_sum = sum(rounded_values.values())
                    diff = 8 - rounded_sum
                    if abs(diff) > 0.01 and rounded_values:
                        # Adjust the largest value
                        max_cc = max(rounded_values, key=lambda k: rounded_values[k])
                        rounded_values[max_cc] = rounding(rounded_values[max_cc] + diff)
                    for cc in agg:
                        if cc not in normalized_agg:
                            normalized_agg[cc] = {}
                        normalized_agg[cc][d] = rounded_values.get(cc, 0)
            
            # Build columns: M-F dates
            columns = ["Cost Center Name"] + [d.strftime("%a %m/%d") for d in dates]
            # Create Treeview
            week_frame = self.week_tables[week_idx]
            for widget in week_frame.winfo_children():
                widget.destroy()
            tree = ttk.Treeview(week_frame, show="headings", columns=columns, height=8)
            tree.grid(row=1, column=0, sticky="ew", columnspan=2, padx=40)
            
            scroll = ttk.Scrollbar(week_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scroll.set)
            scroll.grid(row=1, column=1, sticky="nse")
            for i, col in enumerate(columns):
                if i == 0:
                    tree.column(col, anchor="center", width=180, stretch=True)  # Cost Center Name wider
                else:
                    tree.column(col, anchor="center", width=100, stretch=True)  # Other columns equal
                tree.heading(col, text=col)
            # Insert rows with alternating colors
            for idx, (cost_center, date_dict) in enumerate(sorted(normalized_agg.items())):
                values = [cost_center] + [str(date_dict.get(d, 0)) for d in dates]
                tag = "evenrow" if idx % 2 == 0 else "oddrow"
                tree.insert("", "end", values=values, tags=(tag,))
                tree.tag_configure("evenrow", background="#e6f2ff")  # light blue
                tree.tag_configure("oddrow", background="#ffffff")    # white

            # Calculate total actual hours for each day (not normalized)
            total_actuals = []
            for d in dates:
                total_actuals.append(str(round(sum(agg.get(cc, {}).get(d, 0) for cc in agg), 2)))

            # Insert the total row
            tree.insert("", "end", values=["Total"] + total_actuals, tags=("totalrow",))
            tree.tag_configure("totalrow", background="#accda3")  # white
            # Add week label
            # If this is the current week, say 'Current Week' instead
            today_date = self.today.date() if hasattr(self.today, 'date') else self.today
            is_current_week = today_date >= dates[0] and today_date <= dates[-1]
            if is_current_week:
                label_text = f"Current Week ({dates[0].strftime('%b %d, %Y')})"
            else:
                label_text = f"Week of {dates[0].strftime('%b %d, %Y')}"
            week_label = tk.Label(week_frame, text=label_text, font=self.font_header)
            week_label.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(5,0))

    


    def on_tab_changed(self, event):
        selected_tab = event.widget.select()
        tab_text = event.widget.tab(selected_tab, "text")
        if tab_text == "View Hrs":
            self.fetch_excel_data()
            self.display_timesheet_data()
        
    def is_valid_positive_number(self, value):
        if value == "":
            return True
        if value.startswith("."):
            value = "0" + value
        try:
            val = float(value)
            return val >= 0
        except ValueError:
            return False
        

    def tab_config(self):
        self.tab_config_frame = tk.Frame(self.notebook, bg="lightblue")
        self.notebook.add(self.tab_config_frame, text="Config")

        self.tab_config_frame.columnconfigure(1, weight=1)

        r=0
        c=0
        tk.Label(self.tab_config_frame, text="Configure!", font=self.font_header).grid(row=r, column=c, sticky="ew", padx=10, pady=30)
        self.locations_entry = tk.Entry(self.tab_config_frame, width=50)

        # Error label for file issues
        self.config_error_label = tk.Label(self.tab_config_frame, text="", font=self.font_normal, fg="red")
        self.config_error_label.grid(row=r+10, column=c, columnspan=3, sticky="ew", padx=10, pady=10)

        # Locations YAML
        tk.Label(self.tab_config_frame, text="Locations YAML:", font=self.font_normal).grid(row=r+1, column=c, sticky="ns", padx=10, pady=10)
        self.locations_entry = tk.Entry(self.tab_config_frame, width=50)
        self.locations_entry.insert(0, self.locations_yaml_path)
        self.locations_entry.grid(row=r+1, column=c+1, sticky="nsew",padx=10, pady=10)
        tk.Button(self.tab_config_frame, text="Browse", command=self.browse_locations_yaml).grid(row=r+1, column=c+2, sticky="ns",padx=10, pady=10)

        # Projects YAML
        tk.Label(self.tab_config_frame, text="Projects YAML:", font=self.font_normal).grid(row=r+2, column=c, sticky="ns", padx=10, pady=10)
        self.projects_entry = tk.Entry(self.tab_config_frame, width=50)
        self.projects_entry.insert(0, self.projects_yaml_path)
        self.projects_entry.grid(row=r+2, column=c+1, sticky="nsew",padx=10, pady=10)
        tk.Button(self.tab_config_frame, text="Browse", command=self.browse_projects_yaml).grid(row=r+2, column=c+2, sticky="ns",padx=10, pady=10)

        # Timesheet Excel
        tk.Label(self.tab_config_frame, text="Timesheet Excel:", font=self.font_normal).grid(row=r+3, column=c, sticky="ns", padx=10, pady=10)
        self.excel_entry = tk.Entry(self.tab_config_frame, width=50)
        self.excel_entry.insert(0, self.timesheet_excel_path)
        self.excel_entry.grid(row=r+3, column=c+1, sticky="nsew",padx=10, pady=10)
        tk.Button(self.tab_config_frame, text="Browse", command=self.browse_timesheet_excel).grid(row=r+3, column=c+2, sticky="ns",padx=10, pady=10)

        self.check_config_files()


    def check_config_files(self):
        error_msgs = []
        # Check locations YAML
        if not os.path.exists(self.locations_yaml_path):
            error_msgs.append("Locations YAML file does not exist.")
        # Check projects YAML
        if not os.path.exists(self.projects_yaml_path):
            error_msgs.append("Projects YAML file does not exist.")
        # Check Excel file
        if not os.path.exists(self.timesheet_excel_path):
            error_msgs.append("Timesheet Excel file does not exist.")
        # Try opening Excel file
        try:
            if os.path.exists(self.timesheet_excel_path):
                load_workbook(self.timesheet_excel_path)
        except Exception:
            error_msgs.append("Timesheet Excel file cannot be opened (may be open in another program or corrupted).")
        # Show error message
        if error_msgs:
            self.config_error_label.config(text="\n".join(error_msgs))
        else:
            self.config_error_label.config(text="")

    def browse_locations_yaml(self):
        path = filedialog.askopenfilename(title="Select Locations YAML", filetypes=[("YAML files", "*.yaml")])
        if path:
            self.locations_entry.delete(0, tk.END)
            self.locations_entry.insert(0, path)
            self.locations_yaml_path = path
            self.save_config()
            self.dict_locations = self.import_yaml(self.locations_yaml_path)
            self.clear_notebook_tabs()
            self.tab_input()
            self.tab_display()
            self.tab_config()
            self.check_config_files()
            self.notebook.select(self.notebook.tabs()[-1])

    def browse_projects_yaml(self):
        path = filedialog.askopenfilename(title="Select Projects YAML", filetypes=[("YAML files", "*.yaml")])
        if path:
            self.projects_entry.delete(0, tk.END)
            self.projects_entry.insert(0, path)
            self.projects_yaml_path = path
            self.save_config()
            self.dict_projects = self.import_yaml(self.projects_yaml_path)
            if "Retired" in self.dict_projects:
                self.dict_projects.pop("Retired")
            self.clear_notebook_tabs()
            self.tab_input()
            self.tab_display()
            self.tab_config()
            self.check_config_files()
            self.notebook.select(self.notebook.tabs()[-1])

    def browse_timesheet_excel(self):
        path = filedialog.askopenfilename(title="Select Timesheet Excel", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.excel_entry.delete(0, tk.END)
            self.excel_entry.insert(0, path)
            self.timesheet_excel_path = path
            self.save_config()
            self.fetch_excel_data()
            self.clear_notebook_tabs()
            self.tab_input()
            self.tab_display()
            self.tab_config()
            self.check_config_files()
            self.notebook.select(self.notebook.tabs()[-1])

    def clear_notebook_tabs(self):
        for tab_id in self.notebook.tabs():
            self.notebook.forget(tab_id)


    def tab_monthly_fte(self):
        tab_fte = tk.Frame(self.notebook)
        
        self.notebook.add(tab_fte, text="Monthly FTE")
        tab_fte.columnconfigure(0, weight=1)
        tab_fte.rowconfigure(0, weight=1)

        # Create canvas and scrollbar for scrolling
        canvas = tk.Canvas(tab_fte)
        scrollbar = ttk.Scrollbar(tab_fte, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        fte_frame = tk.Frame(canvas)
        fte_frame.columnconfigure(0, weight=1)
        fte_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=fte_frame, anchor="nw")

        # Gather all data from Excel
        all_data = []
        for row in self.timesheet_ws_hrs.iter_rows(min_row=2, values_only=True):
            date = row[self.timesheet_ws_hrs_column_index["date"] - 1]
            project_nickname = row[self.timesheet_ws_hrs_column_index["project_nickname"] - 1]
            hrs = row[self.timesheet_ws_hrs_column_index["hrs"] - 1]
            if date and project_nickname and hrs:
                try:
                    hrs_val = float(hrs)
                except ValueError:
                    continue
                # Convert date to datetime.date
                if isinstance(date, str):
                    try:
                        date_obj = pd.to_datetime(date).date()
                    except Exception:
                        continue
                else:
                    date_obj = date if hasattr(date, 'date') else date
                all_data.append((date_obj, project_nickname, hrs_val))

        # Group by month
        
        month_project_hrs = defaultdict(lambda: defaultdict(float))
        month_total_hrs = defaultdict(float)
        for date_obj, project_nickname, hrs_val in all_data:
            month = date_obj.strftime('%Y-%m')
            month_project_hrs[month][project_nickname] += hrs_val
            month_total_hrs[month] += hrs_val

        # Sort months chronologically
        months_sorted = reversed(sorted(month_project_hrs.keys()))

        for month_idx, month in enumerate(months_sorted):
            month_dt = datetime.strptime(month, '%Y-%m')
            month_label = tk.Label(fte_frame, text=f"{month_dt.strftime('%b %Y')}", font=self.font_header)
            month_label.grid(row=month_idx*2, column=0, columnspan=3, sticky="ew", pady=(10,0))

            columns = ["Project Nickname", "Total Hrs", "% of Month"]
            tree = ttk.Treeview(fte_frame, show="headings", columns=columns, height=10)
            tree.grid(row=month_idx*2+1, column=0, columnspan=3, sticky="ew", padx=40)
            for i, col in enumerate(columns):
                if i == 0:
                    tree.column(col, anchor="center", width=180, stretch=True)
                else:
                    tree.column(col, anchor="center", width=100, stretch=True)
                tree.heading(col, text=col)

            # Insert rows
            total_hrs = month_total_hrs[month]
            for idx, (project_nickname, hrs) in enumerate(sorted(month_project_hrs[month].items())):
                percent = (hrs / total_hrs * 100) if total_hrs > 0 else 0
                hrs_rounded = round(hrs)
                percent_rounded = round(percent)
                values = [project_nickname, f"{hrs_rounded}", f"{percent_rounded}%"]
                tag = "evenrow" if idx % 2 == 0 else "oddrow"
                tree.insert("", "end", values=values, tags=(tag,))
                tree.tag_configure("evenrow", background="#e6f2ff")
                tree.tag_configure("oddrow", background="#ffffff")

            # Insert total row
            tree.insert("", "end", values=["Total", f"{round(total_hrs)}", "100%"], tags=("totalrow",))
            tree.tag_configure("totalrow", background="#accda3")

    def tab_yearly_fte(self):
        tab_year = tk.Frame(self.notebook)
        self.notebook.add(tab_year, text="Yearly FTE")
        tab_year.columnconfigure(0, weight=1)
        tab_year.rowconfigure(0, weight=1)

        # Create canvas and scrollbar for scrolling
        canvas = tk.Canvas(tab_year)
        scrollbar = ttk.Scrollbar(tab_year, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        year_frame = tk.Frame(canvas)
        year_frame.columnconfigure(0, weight=1)
        year_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=year_frame, anchor="nw")

        # Gather all data from Excel
        all_data = []
        for row in self.timesheet_ws_hrs.iter_rows(min_row=2, values_only=True):
            date = row[self.timesheet_ws_hrs_column_index["date"] - 1]
            project_nickname = row[self.timesheet_ws_hrs_column_index["project_nickname"] - 1]
            hrs = row[self.timesheet_ws_hrs_column_index["hrs"] - 1]
            if date and project_nickname and hrs:
                try:
                    hrs_val = float(hrs)
                except ValueError:
                    continue
                # Convert date to datetime.date
                if isinstance(date, str):
                    try:
                        date_obj = pd.to_datetime(date).date()
                    except Exception:
                        continue
                else:
                    date_obj = date if hasattr(date, 'date') else date
                all_data.append((date_obj, project_nickname, hrs_val))

        # Filter for current year
        current_year = self.today.year
        year_project_hrs = defaultdict(float)
        year_total_hrs = 0.0
        for date_obj, project_nickname, hrs_val in all_data:
            if date_obj.year == current_year:
                year_project_hrs[project_nickname] += hrs_val
                year_total_hrs += hrs_val

        # Table label
        year_label = tk.Label(year_frame, text=f"Year: {current_year}", font=self.font_header)
        year_label.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(10,0))

        columns = ["Project Nickname", "Total Hrs", "% of Year"]
        tree = ttk.Treeview(year_frame, show="headings", columns=columns, height=20)
        tree.grid(row=1, column=0, columnspan=3, sticky="ew", padx=40)
        for i, col in enumerate(columns):
            if i == 0:
                tree.column(col, anchor="center", width=180, stretch=True)
            else:
                tree.column(col, anchor="center", width=100, stretch=True)
            tree.heading(col, text=col)

        # Insert rows
        for idx, (project_nickname, hrs) in enumerate(sorted(year_project_hrs.items())):
            percent = (hrs / year_total_hrs * 100) if year_total_hrs > 0 else 0
            hrs_rounded = round(hrs)
            percent_rounded = round(percent)
            values = [project_nickname, f"{hrs_rounded}", f"{percent_rounded}%"]
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            tree.insert("", "end", values=values, tags=(tag,))
            tree.tag_configure("evenrow", background="#e6f2ff")
            tree.tag_configure("oddrow", background="#ffffff")

        # Insert total row
        tree.insert("", "end", values=["Total", f"{round(year_total_hrs)}", "100%"], tags=("totalrow",))
        tree.tag_configure("totalrow", background="#accda3")


        
def rounding(x):
    return round(x,2)

if __name__ == "__main__":
    app = Timesheet()
    app.mainloop()